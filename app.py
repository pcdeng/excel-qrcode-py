# coding=utf-8

from openpyxl import Workbook, load_workbook
import qrcode
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

c2e = cm_to_EMU
p2e = pixels_to_EMU

column_width = 13.25
row_height = 100
xlsxFileName = 'rooms.xlsx'


def generateQrCode(roomName):
    if roomName:
        img = qrcode.make(roomName)
        img_name = "./qrcodes/" + roomName + ".png"
        with open(img_name, 'wb') as f:
            img.save(f)
        return img_name


def cellw(x): return c2e((x * (column_width-1.71))/10)
def cellh(x): return c2e((x * row_height)/99)


def batchGenerateQrcodes():
    wb = load_workbook(xlsxFileName)
    sheet = wb.active
    rowNum = sheet.max_row
    colNum = sheet.max_column
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Calculated number of cells width or height from cm into EMUs

    for i in range(2, rowNum + 1):
        if i > 0:
            for j in range(1, colNum):
                roomName = sheet.cell(row=i, column=j).value
                imageName = generateQrCode(roomName)
                if imageName:
                    cell = 'B' + bytes(i)
                    sheet.column_dimensions['B'].width = column_width
                    sheet.row_dimensions[i].height = row_height
                    sheet[cell].alignment = align

                    img = Image(imageName)
                    newsize = (90, 90)
                    img.width, img.height = newsize
                    column = 1
                    coloffset = cellw(0.05)
                    row = i - 1
                    rowoffset = cellh(0.5)
                    h, w = img.height, img.width
                    size = XDRPositiveSize2D(p2e(h), p2e(w))
                    marker = AnchorMarker(
                        col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                    sheet.add_image(img)

    wb.save(xlsxFileName)


batchGenerateQrcodes()
